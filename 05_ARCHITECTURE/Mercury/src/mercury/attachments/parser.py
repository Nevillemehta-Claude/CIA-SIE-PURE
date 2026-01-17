"""
Mercury File Parser
===================

Parses various file formats into structured data for AI analysis.

Supported Formats:
- CSV: Comma-separated values
- Excel: .xlsx and .xls spreadsheets
- JSON: Structured data
- Text: Plain text files
- Images: PNG, JPG for chart/document analysis
"""

import base64
import csv
import io
import json
from dataclasses import dataclass, field
from datetime import datetime, timezone
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from mercury.core.logging import get_logger

logger = get_logger("mercury.attachments.parser")


class FileType(Enum):
    """Supported file types."""
    CSV = "csv"
    EXCEL = "excel"
    JSON = "json"
    TEXT = "text"
    IMAGE = "image"
    UNKNOWN = "unknown"


@dataclass
class ParsedFile:
    """Result of parsing a file."""
    filename: str
    file_type: FileType
    size_bytes: int
    parsed_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    
    # For tabular data (CSV, Excel)
    headers: List[str] = field(default_factory=list)
    rows: List[List[Any]] = field(default_factory=list)
    row_count: int = 0
    column_count: int = 0
    
    # For JSON data
    json_data: Optional[Dict[str, Any]] = None
    
    # For text data
    text_content: Optional[str] = None
    
    # For images (base64 encoded)
    image_base64: Optional[str] = None
    image_mime_type: Optional[str] = None
    
    # Summary for AI context
    summary: str = ""
    
    # Parsing errors/warnings
    warnings: List[str] = field(default_factory=list)
    
    def to_ai_context(self) -> str:
        """
        Convert parsed file to context string for AI.
        
        Returns a formatted string that can be included in the AI prompt.
        """
        context_parts = [f"## Attached File: {self.filename}"]
        context_parts.append(f"Type: {self.file_type.value}")
        context_parts.append(f"Size: {self._format_size(self.size_bytes)}")
        
        if self.file_type in (FileType.CSV, FileType.EXCEL):
            context_parts.append(f"Rows: {self.row_count}, Columns: {self.column_count}")
            context_parts.append("")
            context_parts.append("### Data Preview")
            context_parts.append(self._format_table_preview())
            
            if self.summary:
                context_parts.append("")
                context_parts.append(f"### Summary: {self.summary}")
        
        elif self.file_type == FileType.JSON:
            context_parts.append("")
            context_parts.append("### JSON Content")
            context_parts.append("```json")
            context_parts.append(json.dumps(self.json_data, indent=2, default=str)[:2000])
            if len(json.dumps(self.json_data, default=str)) > 2000:
                context_parts.append("... (truncated)")
            context_parts.append("```")
        
        elif self.file_type == FileType.TEXT:
            context_parts.append("")
            context_parts.append("### Text Content")
            content = self.text_content[:3000] if self.text_content else ""
            context_parts.append(content)
            if self.text_content and len(self.text_content) > 3000:
                context_parts.append("... (truncated)")
        
        elif self.file_type == FileType.IMAGE:
            context_parts.append("")
            context_parts.append("### Image Attached")
            context_parts.append("(Image data will be sent to AI for visual analysis)")
        
        if self.warnings:
            context_parts.append("")
            context_parts.append("### Warnings")
            for warning in self.warnings:
                context_parts.append(f"- {warning}")
        
        return "\n".join(context_parts)
    
    def _format_table_preview(self, max_rows: int = 10) -> str:
        """Format tabular data as markdown table."""
        if not self.headers:
            return "(No data)"
        
        lines = []
        
        # Header row
        lines.append("| " + " | ".join(str(h)[:20] for h in self.headers) + " |")
        lines.append("| " + " | ".join("---" for _ in self.headers) + " |")
        
        # Data rows (limit to max_rows)
        for row in self.rows[:max_rows]:
            cells = [str(cell)[:20] if cell is not None else "" for cell in row]
            lines.append("| " + " | ".join(cells) + " |")
        
        if self.row_count > max_rows:
            lines.append(f"... and {self.row_count - max_rows} more rows")
        
        return "\n".join(lines)
    
    def _format_size(self, bytes: int) -> str:
        """Format file size in human-readable format."""
        for unit in ['B', 'KB', 'MB', 'GB']:
            if bytes < 1024:
                return f"{bytes:.1f} {unit}"
            bytes /= 1024
        return f"{bytes:.1f} TB"
    
    def get_statistics(self) -> Dict[str, Any]:
        """Get basic statistics for numerical columns."""
        if self.file_type not in (FileType.CSV, FileType.EXCEL):
            return {}
        
        stats = {}
        for col_idx, header in enumerate(self.headers):
            values = []
            for row in self.rows:
                if col_idx < len(row):
                    try:
                        val = float(row[col_idx])
                        values.append(val)
                    except (ValueError, TypeError):
                        pass
            
            if values:
                stats[header] = {
                    "count": len(values),
                    "min": min(values),
                    "max": max(values),
                    "mean": sum(values) / len(values),
                    "sum": sum(values),
                }
        
        return stats


class FileParser:
    """
    Parses files into structured data for AI analysis.
    
    Usage:
        parser = FileParser()
        result = await parser.parse(file_content, filename)
    """
    
    # Maximum file size (10 MB)
    MAX_FILE_SIZE = 10 * 1024 * 1024
    
    # Maximum rows to keep in memory
    MAX_ROWS = 10000
    
    SUPPORTED_EXTENSIONS = {
        ".csv": FileType.CSV,
        ".xlsx": FileType.EXCEL,
        ".xls": FileType.EXCEL,
        ".json": FileType.JSON,
        ".txt": FileType.TEXT,
        ".md": FileType.TEXT,
        ".png": FileType.IMAGE,
        ".jpg": FileType.IMAGE,
        ".jpeg": FileType.IMAGE,
        ".gif": FileType.IMAGE,
        ".webp": FileType.IMAGE,
    }
    
    MIME_TYPES = {
        "text/csv": FileType.CSV,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": FileType.EXCEL,
        "application/vnd.ms-excel": FileType.EXCEL,
        "application/json": FileType.JSON,
        "text/plain": FileType.TEXT,
        "image/png": FileType.IMAGE,
        "image/jpeg": FileType.IMAGE,
        "image/gif": FileType.IMAGE,
        "image/webp": FileType.IMAGE,
    }
    
    def __init__(self):
        self._openpyxl_available = self._check_openpyxl()
    
    def _check_openpyxl(self) -> bool:
        """Check if openpyxl is available for Excel parsing."""
        try:
            import openpyxl
            return True
        except ImportError:
            return False
    
    async def parse(
        self,
        content: bytes,
        filename: str,
        mime_type: Optional[str] = None,
    ) -> ParsedFile:
        """
        Parse file content into structured data.
        
        Args:
            content: Raw file bytes
            filename: Original filename
            mime_type: Optional MIME type hint
        
        Returns:
            ParsedFile with extracted data
        """
        size = len(content)
        
        # Determine file type
        file_type = self._detect_type(filename, mime_type)
        
        result = ParsedFile(
            filename=filename,
            file_type=file_type,
            size_bytes=size,
        )
        
        # Validate size
        if size > self.MAX_FILE_SIZE:
            result.warnings.append(f"File exceeds maximum size of {self.MAX_FILE_SIZE // (1024*1024)} MB")
            result.summary = "File too large to process"
            return result
        
        # Parse based on type
        try:
            if file_type == FileType.CSV:
                await self._parse_csv(content, result)
            elif file_type == FileType.EXCEL:
                await self._parse_excel(content, result)
            elif file_type == FileType.JSON:
                await self._parse_json(content, result)
            elif file_type == FileType.TEXT:
                await self._parse_text(content, result)
            elif file_type == FileType.IMAGE:
                await self._parse_image(content, result, mime_type)
            else:
                result.warnings.append(f"Unsupported file type: {file_type.value}")
        
        except Exception as e:
            logger.error(f"Error parsing {filename}: {e}")
            result.warnings.append(f"Parse error: {str(e)}")
        
        return result
    
    def _detect_type(
        self,
        filename: str,
        mime_type: Optional[str],
    ) -> FileType:
        """Detect file type from extension or MIME type."""
        # Try extension first
        ext = Path(filename).suffix.lower()
        if ext in self.SUPPORTED_EXTENSIONS:
            return self.SUPPORTED_EXTENSIONS[ext]
        
        # Try MIME type
        if mime_type and mime_type in self.MIME_TYPES:
            return self.MIME_TYPES[mime_type]
        
        return FileType.UNKNOWN
    
    async def _parse_csv(self, content: bytes, result: ParsedFile) -> None:
        """Parse CSV file."""
        # Try to decode with different encodings
        text = None
        for encoding in ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']:
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        
        if text is None:
            result.warnings.append("Could not decode file - tried multiple encodings")
            return
        
        # Parse CSV
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        
        if not rows:
            result.warnings.append("CSV file is empty")
            return
        
        # First row as headers
        result.headers = rows[0]
        result.rows = rows[1:self.MAX_ROWS + 1]
        result.row_count = len(rows) - 1
        result.column_count = len(result.headers)
        
        if len(rows) - 1 > self.MAX_ROWS:
            result.warnings.append(f"Data truncated to {self.MAX_ROWS} rows")
        
        # Generate summary
        result.summary = f"CSV with {result.row_count} rows and {result.column_count} columns"
    
    async def _parse_excel(self, content: bytes, result: ParsedFile) -> None:
        """Parse Excel file."""
        if not self._openpyxl_available:
            result.warnings.append("Excel parsing requires 'openpyxl' package")
            result.summary = "Excel file (openpyxl not installed)"
            return
        
        import openpyxl
        
        # Load workbook
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = wb.active
        
        if sheet is None:
            result.warnings.append("No active sheet in workbook")
            return
        
        # Read data
        rows = []
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_idx >= self.MAX_ROWS + 1:
                break
            rows.append(list(row))
        
        if not rows:
            result.warnings.append("Excel sheet is empty")
            return
        
        # First row as headers
        result.headers = [str(h) if h else f"Column_{i}" for i, h in enumerate(rows[0])]
        result.rows = rows[1:]
        result.row_count = sheet.max_row - 1 if sheet.max_row else 0
        result.column_count = len(result.headers)
        
        if result.row_count > self.MAX_ROWS:
            result.warnings.append(f"Data truncated to {self.MAX_ROWS} rows")
        
        result.summary = f"Excel with {result.row_count} rows and {result.column_count} columns"
        
        wb.close()
    
    async def _parse_json(self, content: bytes, result: ParsedFile) -> None:
        """Parse JSON file."""
        text = content.decode('utf-8')
        result.json_data = json.loads(text)
        
        # Summarize structure
        if isinstance(result.json_data, list):
            result.summary = f"JSON array with {len(result.json_data)} items"
        elif isinstance(result.json_data, dict):
            result.summary = f"JSON object with {len(result.json_data)} keys"
        else:
            result.summary = f"JSON value: {type(result.json_data).__name__}"
    
    async def _parse_text(self, content: bytes, result: ParsedFile) -> None:
        """Parse text file."""
        for encoding in ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']:
            try:
                result.text_content = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        
        if result.text_content:
            lines = result.text_content.count('\n') + 1
            words = len(result.text_content.split())
            result.summary = f"Text file with {lines} lines and ~{words} words"
        else:
            result.warnings.append("Could not decode text file")
    
    async def _parse_image(
        self,
        content: bytes,
        result: ParsedFile,
        mime_type: Optional[str],
    ) -> None:
        """Parse image file for AI vision analysis."""
        result.image_base64 = base64.b64encode(content).decode('utf-8')
        
        # Determine MIME type
        if mime_type:
            result.image_mime_type = mime_type
        else:
            # Detect from magic bytes
            if content[:8] == b'\x89PNG\r\n\x1a\n':
                result.image_mime_type = "image/png"
            elif content[:2] == b'\xff\xd8':
                result.image_mime_type = "image/jpeg"
            elif content[:6] in (b'GIF87a', b'GIF89a'):
                result.image_mime_type = "image/gif"
            else:
                result.image_mime_type = "image/png"  # Default
        
        result.summary = f"Image ({result.image_mime_type})"


# Global parser instance
_parser: Optional[FileParser] = None


def get_parser() -> FileParser:
    """Get the global file parser."""
    global _parser
    if _parser is None:
        _parser = FileParser()
    return _parser


async def parse_file(
    content: bytes,
    filename: str,
    mime_type: Optional[str] = None,
) -> ParsedFile:
    """Parse a file (convenience function)."""
    parser = get_parser()
    return await parser.parse(content, filename, mime_type)


def get_supported_formats() -> Dict[str, str]:
    """Get dictionary of supported formats."""
    return {
        "CSV": ".csv - Comma-separated values",
        "Excel": ".xlsx, .xls - Microsoft Excel spreadsheets",
        "JSON": ".json - Structured data",
        "Text": ".txt, .md - Plain text files",
        "Images": ".png, .jpg, .jpeg, .gif, .webp - For chart/document analysis",
    }
